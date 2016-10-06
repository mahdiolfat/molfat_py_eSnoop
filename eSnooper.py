from google import search
from openpyxl import load_workbook
import urllib2
import httplib
import sys
import re
import ssl
import signal

FNAMEIDX = 1
LNAMEIDX = 2
TITLEIDX = 3
AFFILIATIONIDX = 4
ADDRESSIDX = 5
ADDRESS2IDX = 6
CITYIDX = 7
PROVINCEIDX = 8
ZIPIDX = 9
COUNTRYIDX = 10
EMAILIDX = 11

FILE = 'M&M Pre Attendee List 6_30.xlsx'

STARTIDX = 40
ENDIDX = 658

def signal_handler(signal, frame):
    print('Saving Document')
    es.wb.save(FILE)
    sys.exit(0)

signal.signal(signal.SIGINT, signal_handler)

def StripTags(text):
    finished = 0
    while not finished:
        finished = 1
        start = text.find("<")
        if start >= 0:
            stop = text[start:].find(">")
            if stop >= 0:
                text = text[:start] + text[start+stop+1:]
                finished = 0
    return text

class Person:
    def __init__(self,  idx,
                        fName = '',
                        lName = '',
                        title = '',
                        affiliation = '',
                        address = '',
                        country = '',
                        province = ''):
        self.fName = fName
        self.lName = lName
        self.title = title
        self.affiliation = affiliation
        self.country = country
        self.province = province
        self.idx = idx

    def printInfo(self):
        print "Name: " + self.fName + " " + self.lName
        print "Title: " + self.title
        print "Affiliation: " + self.affiliation
        print "Country: " + self.country
        print

class EmailSnooper:
    def __init__(self, workbook, address = ''):
        self.count = 0
        self.wb = load_workbook(filename = address + workbook)
        self.ws = self.wb.active
        self.persons = []
        self.emails = []
        print self.ws

    def loadAllPersons(self):
        for i in range(STARTIDX, ENDIDX):
            p = Person(idx = i)
            self.persons.append(p)
            p.fName = self.ws.cell(row = i, column = FNAMEIDX).value
            p.lName = self.ws.cell(row = i, column = LNAMEIDX).value
            p.title = self.ws.cell(row = i, column = TITLEIDX).value
            p.affiliation = self.ws.cell(row = i, column = AFFILIATIONIDX).value
            p.address = self.ws.cell(row = i, column = ADDRESSIDX).value
            p.country = self.ws.cell(row = i, column = COUNTRYIDX).value
            p.province = self.ws.cell(row = i, column = PROVINCEIDX).value

    def printPersons(self):
        for person in self.persons:
            person.printInfo()

    def queryTxt(self, person):
        query = ""
        if person.fName is not None:
            query += person.fName
        if person.lName is not None:
            query += " " + person.lName
        if person.affiliation is not None:
            query += " " + person.affiliation
        #if person.title is not None:
        #    query += " " + person.title

        return query

    def scrapeAllPersons(self):
        for person in self.persons:
            emails_str = ""
            print
            emails = self.scrapePerson(person)

            if len(emails) is 0:
                continue #move on to next person

            person.emails = emails
            for email in emails:
                emails_str += email + ", "

            print "Writing emails"
            self.ws.cell(row = person.idx, column = EMAILIDX).value = emails_str

    def scrapePerson(self, person):
        print type(person)
        if not isinstance(person, Person):
            print "Wrong person object provided"
            return

        print "Email for " + person.fName + " " + person.lName
        # conduct a google search
        d = {}
        try:
            for url in search(self.queryTxt(person).encode("utf-8"), stop=10):
                try:
                    request = urllib2.Request(url)
                    request.add_header('User-Agent', 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/39.0.2171.71 Safari/537.36')
                    #response = urllib2.urlopen(request)
                    opener = urllib2.build_opener()
                    response = opener.open(request)
                    try:
                        text = unicode(response.read(), "utf-8", errors='ignore')
                    except UnicodeError as e:
                        print e
                        print "FUUUUUCK"
                    except KeyError as e:
                        print e

                    #print text
                    emails = re.findall(r'\w+[.|\w]\w+@\w+[.]\w+[.|\w+]\w+', text)
                    for email in emails:
                        d[email] = 1
                        #print "EMAIL FOUND: " + email
                    #print len(emails)
                    response.close()
                except IOError as e:
                    print e
                except urllib2.URLError as e:
                    print e
                except Exception as e:
                    print e
                    self.wb.save(FILE)
        except KeyError as e:
            print e

        return d.keys()

es = EmailSnooper(workbook = FILE)
if __name__ == "__main__":
    es.loadAllPersons()

    # print all persons' information
    #es.printPersons()

    #es.scrapePerson(es.persons[2])
    es.scrapeAllPersons()

    #print es.persons[0].emails
    es.wb.save(FILE)

